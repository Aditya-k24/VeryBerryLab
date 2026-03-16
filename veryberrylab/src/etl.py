"""
src/etl.py
==========
Ingest Phenotyping 4 Worksheet 2 (wide cross-tab, 11 cultivar sheets) and
produce a tidy, row-per-rep CSV with batch_id and missing_reason semantics.

Missing-reason vocabulary
--------------------------
  observed       — numeric value present
  not_measured   — cell was '-' (scheduled date, value not recorded)
  not_scheduled  — this date does not exist for this cultivar (different batch)

Canonical data source
---------------------
    ../Phenotyping Data with Aditya 1_27_2026/Pheno Batch 4/
        Phenotyping 4 Worksheet 2.xlsx

Run from the veryberrylab/ directory:
    python3 src/etl.py

Output
------
    data/processed/pheno4_clean.csv
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent

CANONICAL_XLSX = (
    ROOT.parent
    / "Phenotyping Data with Aditya 1_27_2026"
    / "Pheno Batch 4"
    / "Phenotyping 4 Worksheet 2.xlsx"
)
FALLBACK_XLSX = ROOT / "data" / "raw" / "Phenotyping 4 Worksheet 2.xlsx"
RAW_PATH = CANONICAL_XLSX if CANONICAL_XLSX.exists() else FALLBACK_XLSX

OUT_PATH = ROOT / "data" / "processed" / "pheno4_clean.csv"
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Cultivar assignment
# ---------------------------------------------------------------------------
CULTIVAR_MAP = {
    "Fin":   "Finn",
    "Cab":   "Cabrio",
    "Cam":   "Camarosa",   # provisional — confirm with team
    "Sen":   "Sensation",
    "Cha":   "Chandler",
    "Alb ":  "Albion",
    "Alb":   "Albion",
    "Mox":   "Moxie",
    "RJune": "Ruby June",
    "Bri":   "Brilliance",
    "Por":   "Portola",
    "Rad":   "Radiance",
}

BATCH_A = {"Albion", "Cabrio", "Camarosa", "Chandler", "Finn", "Sensation"}
BATCH_B = {"Brilliance", "Moxie", "Portola", "Radiance", "Ruby June"}

# ---------------------------------------------------------------------------
# Trait map
# ---------------------------------------------------------------------------
TRAIT_MAP = {
    "Pri stolon":                "n_stolon_primary",
    "Sec stolon":                "n_stolon_secondary",
    "Ter stolon":                "n_stolon_tertiary",
    "Quart Stolon":              "n_stolon_quaternary",
    "dp on alt of pri stolon":   "n_dp_alt_primary",
    "dp on alt of sec stolon":   "n_dp_alt_secondary",
    "dp on alt of ter stolon":   "n_dp_alt_tertiary",
    "dp on alt of quart stolon": "n_dp_alt_quaternary",
    "dp on mid of pri stolon":   "n_dp_mid_primary",
    "dp on mid of sec stolon":   "n_dp_mid_secondary",
    "dp on mid of ter stolon":   "n_dp_mid_tertiary",
    "dp on mid of quart stolon": "n_dp_mid_quaternary",
    "Total dp on alt":           "n_dp_total_alt",
    "Total dp on mid":           "n_dp_total_mid",
    "#Total Flowers":            "n_flowers_total",
    "# Flowers mp/mp":           "n_flowers_mp",
    "# Flowers dp/mp":           "n_flowers_dp",
    "Pri Stolon length (cm)":    "stolon_length_primary_cm",
    "Crown diameter (mm)":       "crown_diameter_mm",
}

TRAIT_COLS = list(TRAIT_MAP.values())

TRAIT_LABELS = {
    "n_stolon_primary":          "Primary Stolons",
    "n_stolon_secondary":        "Secondary Stolons",
    "n_stolon_tertiary":         "Tertiary Stolons",
    "n_stolon_quaternary":       "Quaternary Stolons",
    "n_dp_alt_primary":          "Daughter Plants (Alt) — Primary",
    "n_dp_alt_secondary":        "Daughter Plants (Alt) — Secondary",
    "n_dp_alt_tertiary":         "Daughter Plants (Alt) — Tertiary",
    "n_dp_alt_quaternary":       "Daughter Plants (Alt) — Quaternary",
    "n_dp_mid_primary":          "Daughter Plants (Mid) — Primary",
    "n_dp_mid_secondary":        "Daughter Plants (Mid) — Secondary",
    "n_dp_mid_tertiary":         "Daughter Plants (Mid) — Tertiary",
    "n_dp_mid_quaternary":       "Daughter Plants (Mid) — Quaternary",
    "n_dp_total_alt":            "Total DPs — Alternate Nodes",
    "n_dp_total_mid":            "Total DPs — Mid Nodes",
    "n_flowers_total":           "Total Flowers",
    "n_flowers_mp":              "Flowers on Mother Plant",
    "n_flowers_dp":              "Flowers on Daughter Plants",
    "stolon_length_primary_cm":  "Primary Stolon Length (cm)",
    "crown_diameter_mm":         "Crown Diameter (mm)",
}

# Directionality: True = higher is better
TRAIT_DIRECTION = {
    "n_stolon_primary":          True,
    "n_stolon_secondary":        True,
    "n_stolon_tertiary":         True,
    "n_stolon_quaternary":       True,
    "n_dp_alt_primary":          True,
    "n_dp_alt_secondary":        True,
    "n_dp_alt_tertiary":         True,
    "n_dp_alt_quaternary":       True,
    "n_dp_mid_primary":          True,
    "n_dp_mid_secondary":        True,
    "n_dp_mid_tertiary":         True,
    "n_dp_mid_quaternary":       True,
    "n_dp_total_alt":            True,
    "n_dp_total_mid":            True,
    "n_flowers_total":           True,
    "n_flowers_mp":              True,
    "n_flowers_dp":              True,
    "stolon_length_primary_cm":  True,
    "crown_diameter_mm":         True,
}

ALL_CULTIVARS = sorted(BATCH_A | BATCH_B)


# ---------------------------------------------------------------------------
# Per-sheet parser
# ---------------------------------------------------------------------------

def _parse_sheet(ws, cultivar_name: str) -> tuple[list[dict], list[str]]:
    """
    Returns (records, warnings).
    records: list of tidy dicts, one per (date, rep)
    warnings: list of human-readable warning strings
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return [], [f"{cultivar_name}: sheet has fewer than 3 rows"]

    warnings = []
    date_row = rows[0]
    rep_row  = rows[1]

    # Forward-fill dates (merged cells appear as None)
    dates: list = []
    current_date = None
    for val in date_row[1:]:
        if val is not None:
            current_date = val
        dates.append(current_date)

    reps: list = []
    for val in rep_row[1:]:
        try:
            reps.append(int(val))
        except (TypeError, ValueError):
            reps.append(None)

    col_headers = list(zip(dates, reps))

    # Skeleton records: one per (date, rep)
    skeleton: dict[tuple, dict] = {}
    for date, rep in col_headers:
        key = (date, rep)
        if key not in skeleton:
            skeleton[key] = {"cultivar": cultivar_name, "date": date, "rep": rep}

    # Fill traits
    for row in rows[2:]:
        label = row[0]
        if label is None:
            continue
        internal = TRAIT_MAP.get(str(label).strip())
        if internal is None:
            continue

        for (date, rep), raw_val in zip(col_headers, row[1:]):
            key = (date, rep)
            if raw_val == "-" or raw_val == "--" or raw_val is None:
                val = float("nan")
            else:
                try:
                    val = float(raw_val)
                except (TypeError, ValueError):
                    warnings.append(
                        f"{cultivar_name}: unexpected value '{raw_val}' for {internal} — treated as missing"
                    )
                    val = float("nan")
            skeleton[key][internal] = val

    return list(skeleton.values()), warnings


# ---------------------------------------------------------------------------
# Main ETL
# ---------------------------------------------------------------------------

def run_etl(path: Path | None = None, verbose: bool = True) -> tuple[pd.DataFrame, list[str]]:
    """
    Parse the workbook and return (df_clean, ingestion_warnings).
    df_clean has columns: date, cultivar, batch, rep, <trait_cols...>
    """
    source = path or RAW_PATH
    if verbose:
        print(f"Loading: {source}")

    wb = openpyxl.load_workbook(str(source), read_only=True, data_only=True)

    all_records: list[dict] = []
    all_warnings: list[str] = [
        "Formula results rely on cached Excel values. If you see unexpected blanks, "
        "open the file in Excel, recalculate, and re-save before uploading."
    ]

    for tab_name in wb.sheetnames:
        cultivar = CULTIVAR_MAP.get(tab_name)
        if cultivar is None:
            all_warnings.append(f"Unknown sheet tab '{tab_name}' — skipped")
            continue
        ws = wb[tab_name]
        records, sheet_warns = _parse_sheet(ws, cultivar)
        all_records.extend(records)
        all_warnings.extend(sheet_warns)
        if verbose:
            print(f"  [OK]   {tab_name:8s} → {cultivar:12s}  ({len(records)} rows)")

    wb.close()

    df = _build_dataframe(all_records)

    if verbose:
        _validate(df)
        df.to_csv(str(OUT_PATH), index=False)
        print(f"\nSaved → {OUT_PATH}")
        print(f"Shape : {df.shape[0]} rows × {df.shape[1]} columns")

    return df, all_warnings


def _build_dataframe(records: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(records)

    for col in TRAIT_COLS:
        if col not in df.columns:
            df[col] = float("nan")

    df["date"] = pd.to_datetime(df["date"])
    df["rep"]  = pd.to_numeric(df["rep"], errors="coerce").astype("Int64")

    for col in TRAIT_COLS:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # Batch assignment
    df["batch"] = df["cultivar"].apply(
        lambda cv: "A" if cv in BATCH_A else ("B" if cv in BATCH_B else "?")
    )

    col_order = ["date", "cultivar", "batch", "rep"] + TRAIT_COLS
    df = df[[c for c in col_order if c in df.columns]]
    df = df.sort_values(["cultivar", "date", "rep"]).reset_index(drop=True)
    return df


def run_etl_from_bytes(xlsx_bytes: bytes) -> tuple[pd.DataFrame, list[str]]:
    """Parse from raw bytes (e.g. a file upload). Does not write to disk."""
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)

    all_records: list[dict] = []
    all_warnings: list[str] = [
        "Formula results rely on cached Excel values. If you see unexpected blanks, "
        "open the file in Excel, recalculate, and re-save before uploading."
    ]

    for tab_name in wb.sheetnames:
        cultivar = CULTIVAR_MAP.get(tab_name)
        if cultivar is None:
            all_warnings.append(f"Unknown sheet tab '{tab_name}' — skipped")
            continue
        ws = wb[tab_name]
        records, sheet_warns = _parse_sheet(ws, cultivar)
        all_records.extend(records)
        all_warnings.extend(sheet_warns)

    wb.close()

    if not all_records:
        return pd.DataFrame(), all_warnings

    return _build_dataframe(all_records), all_warnings


def _validate(df: pd.DataFrame) -> None:
    print("\n--- Validation ---")
    expected = {
        "Finn", "Cabrio", "Camarosa", "Sensation", "Chandler",
        "Albion", "Moxie", "Ruby June", "Brilliance", "Portola", "Radiance",
    }
    found = set(df["cultivar"].unique())
    if found != expected:
        print(f"  [WARN] Cultivar mismatch. Missing: {expected-found}  Extra: {found-expected}")
    else:
        print(f"  [OK]   All 11 cultivars present")

    bad_reps = df[~df["rep"].isin([1, 2, 3])]
    if not bad_reps.empty:
        print(f"  [FAIL] {len(bad_reps)} rows with unexpected rep IDs")
    else:
        print("  [OK]   All rep IDs are 1, 2, or 3")

    dupes = df.duplicated(["date", "cultivar", "rep"])
    if dupes.any():
        print(f"  [FAIL] {dupes.sum()} duplicate (date, cultivar, rep) rows")
    else:
        print("  [OK]   No duplicate (date, cultivar, rep) rows")

    print(f"  [OK]   Date range: {df['date'].min().date()} → {df['date'].max().date()}")
    print(f"  [OK]   Batch A: {sorted(BATCH_A & found)}  Batch B: {sorted(BATCH_B & found)}")


if __name__ == "__main__":
    df, warnings = run_etl()
    print("\nIngestion warnings:")
    for w in warnings:
        print(f"  • {w}")
    print("\nSample (first 5 rows):")
    pd.set_option("display.max_columns", 10)
    pd.set_option("display.width", 120)
    print(df.head())
