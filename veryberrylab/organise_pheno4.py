"""
veryberrylab/organise_pheno4.py
================================
Reads Phenotyping 4 Worksheet 2.xlsx (wide cross-tab, 11 cultivar sheets)
and writes a clean Pheno 4 Organised.xlsx with one sheet per measurement date.

Each output sheet has:
  Cultivar | Rep | trait | Mean | SD | SE | trait | Mean | SD | SE | ...

Rep 1 rows contain Excel AVERAGE/STDEV/SE formulas referencing the 3 rep rows.
Rep 2 and 3 rows leave Mean/SD/SE blank.

Run from the VeryBerryLab/ directory:
    python3 veryberrylab/organise_pheno4.py
"""

from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT        = Path(__file__).resolve().parent.parent
BATCH4      = ROOT / "Phenotyping Data with Aditya 1_27_2026" / "Pheno Batch 4"
INPUT_PATH  = BATCH4 / "Phenotyping 4 Worksheet 2.xlsx"
OUTPUT_PATH = BATCH4 / "Pheno 4 Organised.xlsx"

# ---------------------------------------------------------------------------
# Cultivar tab → full name  (same as src/etl.py)
# ---------------------------------------------------------------------------
CULTIVAR_MAP = {
    "Fin":   "Finn",
    "Cab":   "Cabrio",
    "Cam":   "Camarosa",
    "Sen":   "Sensation",
    "Cha":   "Chandler",
    "Alb ":  "Albion",     # tab has trailing space
    "Alb":   "Albion",
    "Mox":   "Moxie",
    "RJune": "Ruby June",
    "Bri":   "Brilliance",
    "Por":   "Portola",
    "Rad":   "Radiance",
}

# ---------------------------------------------------------------------------
# Traits: (source row label, internal key, display header)
# Order here determines column order in output.
# ---------------------------------------------------------------------------
TRAITS = [
    ("Pri stolon",                "pri_stolon",   "Pri stolon"),
    ("Sec stolon",                "sec_stolon",   "Sec stolon"),
    ("Ter stolon",                "ter_stolon",   "Ter stolon"),
    ("Quart Stolon",              "quart_stolon", "Quart stolon"),
    ("dp on alt of pri stolon",   "dp_alt_pri",   "dp alt primary"),
    ("dp on alt of sec stolon",   "dp_alt_sec",   "dp alt secondary"),
    ("dp on alt of ter stolon",   "dp_alt_ter",   "dp alt tertiary"),
    ("dp on alt of quart stolon", "dp_alt_quart", "dp alt quaternary"),
    ("dp on mid of pri stolon",   "dp_mid_pri",   "dp mid primary"),
    ("dp on mid of sec stolon",   "dp_mid_sec",   "dp mid secondary"),
    ("dp on mid of ter stolon",   "dp_mid_ter",   "dp mid tertiary"),
    ("dp on mid of quart stolon", "dp_mid_quart", "dp mid quaternary"),
    ("Total dp on alt",           "total_dp_alt", "Total dp alt"),
    ("Total dp on mid",           "total_dp_mid", "Total dp mid"),
    ("#Total Flowers",            "total_flowers","Total flowers"),
    ("# Flowers mp/mp",           "flowers_mp",   "Flowers mp"),
    ("# Flowers dp/mp",           "flowers_dp",   "Flowers dp"),
    ("Pri Stolon length (cm)",    "stolon_len",   "Pri stolon length (cm)"),
    ("Crown diameter (mm)",       "crown_diam",   "Crown diameter (mm)"),
]

LABEL_TO_KEY  = {label: key     for label, key, _       in TRAITS}
TRAIT_KEYS    = [key             for _,     key, _       in TRAITS]
TRAIT_DISPLAY = [display         for _,     _,   display in TRAITS]

# Traits where 0.00 format is appropriate (continuous measurements)
MORPH_KEYS = {"stolon_len", "crown_diam"}

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HEADER_FONT    = Font(bold=True)
HEADER_FILL    = PatternFill("solid", fgColor="C8E6C9")  # light green
ALT_FILL       = PatternFill("solid", fgColor="F5F5F5")  # very light grey
WHITE_FILL     = PatternFill("solid", fgColor="FFFFFF")
THIN           = Side(style="thin")
CENTER         = Alignment(horizontal="center")

# Total data columns: 2 fixed + 19 traits × 4 = 78
TOTAL_COLS = 2 + len(TRAITS) * 4


# ---------------------------------------------------------------------------
# Step 1 — Read raw data
# ---------------------------------------------------------------------------

def read_raw_data(path: Path) -> dict:
    """
    Parse all cultivar sheets from the wide cross-tab workbook.

    Returns nested dict:
        data[date][cultivar][rep][trait_key] = float | None
    """
    data: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    for tab_name in wb.sheetnames:
        cultivar = CULTIVAR_MAP.get(tab_name)
        if cultivar is None:
            print(f"  [SKIP] Unknown tab '{tab_name}' — not in CULTIVAR_MAP")
            continue

        ws = wb[tab_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            print(f"  [SKIP] Tab '{tab_name}' has fewer than 3 rows")
            continue

        date_row = rows[0]   # ('Date', datetime|None, None, None, datetime|None, ...)
        rep_row  = rows[1]   # ('Rep',  1.0, 2.0, 3.0, 1.0, ...)

        # Forward-fill dates across merged-cell Nones
        dates: list = []
        current_date = None
        for val in date_row[1:]:
            if val is not None:
                current_date = val
            dates.append(current_date)

        # Parse rep numbers
        reps: list = []
        for val in rep_row[1:]:
            try:
                reps.append(int(val))
            except (TypeError, ValueError):
                reps.append(None)

        col_headers = list(zip(dates, reps))

        # Parse each trait row
        for row in rows[2:]:
            label = row[0]
            if label is None:
                continue
            key = LABEL_TO_KEY.get(str(label).strip())
            if key is None:
                continue   # not a tracked trait

            for (date, rep), raw_val in zip(col_headers, row[1:]):
                if date is None or rep is None:
                    continue
                if raw_val == "-" or raw_val is None:
                    val = None
                else:
                    try:
                        val = float(raw_val)
                    except (TypeError, ValueError):
                        val = None
                data[date][cultivar][rep][key] = val

        print(f"  [OK]   {tab_name:8s} → {cultivar}")

    wb.close()
    return data


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _sheet_name_for_date(d) -> str:
    """Return 'Apr 16', 'Apr 30', etc. Works with datetime and date objects."""
    month = d.strftime("%b")
    day   = str(d.day)          # no leading zero, portable across platforms
    return f"{month} {day}"


# ---------------------------------------------------------------------------
# Step 3 — Write output workbook
# ---------------------------------------------------------------------------

def write_output(data: dict, output_path: Path) -> None:
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)   # remove default blank sheet

    # Sort dates chronologically
    def _date_key(d):
        if isinstance(d, datetime):
            return d
        # openpyxl may return datetime.date; convert to datetime for comparison
        return datetime(d.year, d.month, d.day)

    all_dates = sorted(data.keys(), key=_date_key)

    for date in all_dates:
        sheet_name      = _sheet_name_for_date(date)
        ws              = wb_out.create_sheet(title=sheet_name)
        cultivars       = sorted(data[date].keys())

        # ---- Header row (row 1) ----------------------------------------
        ws.cell(1, 1).value = "Cultivar"
        ws.cell(1, 2).value = "Rep"
        for c in [1, 2]:
            ws.cell(1, c).font      = HEADER_FONT
            ws.cell(1, c).fill      = HEADER_FILL
            ws.cell(1, c).alignment = CENTER
            ws.cell(1, c).border    = Border(bottom=THIN)

        col = 3
        for display in TRAIT_DISPLAY:
            for label in (display, "Mean", "SD", "SE"):
                cell            = ws.cell(1, col)
                cell.value      = label
                cell.font       = HEADER_FONT
                cell.fill       = HEADER_FILL
                cell.alignment  = CENTER
                cell.border     = Border(bottom=THIN)
                col += 1

        # ---- Data rows --------------------------------------------------
        current_row = 2

        for cult_idx, cultivar in enumerate(cultivars):
            row_fill    = ALT_FILL if cult_idx % 2 == 1 else WHITE_FILL
            rep_data    = data[date][cultivar]
            rep1_row    = current_row   # Excel row number for Rep 1

            for rep in (1, 2, 3):
                trait_vals = rep_data.get(rep, {})

                # Apply background fill to every cell in the row
                for c in range(1, TOTAL_COLS + 1):
                    ws.cell(current_row, c).fill = row_fill

                # Group separator: thin top border on Rep 1 of each cultivar
                # (skip the very first cultivar group)
                if rep == 1 and cult_idx > 0:
                    for c in range(1, TOTAL_COLS + 1):
                        ws.cell(current_row, c).border = Border(top=THIN)

                # Fixed columns
                if rep == 1:
                    ws.cell(current_row, 1).value = cultivar
                ws.cell(current_row, 1).alignment = CENTER
                ws.cell(current_row, 2).value     = rep
                ws.cell(current_row, 2).alignment = CENTER

                # Trait columns
                trait_col = 3
                for key in TRAIT_KEYS:
                    val        = trait_vals.get(key)
                    is_morph   = key in MORPH_KEYS
                    raw_cell   = ws.cell(current_row, trait_col)
                    mean_col   = trait_col + 1
                    sd_col     = trait_col + 2
                    se_col     = trait_col + 3

                    if val is not None:
                        raw_cell.value         = val
                        raw_cell.number_format = "0.00" if is_morph else "0"
                    raw_cell.alignment = CENTER

                    if rep == 1:
                        r1 = rep1_row
                        r3 = rep1_row + 2
                        tc = get_column_letter(trait_col)
                        sc = get_column_letter(sd_col)

                        mean_cell = ws.cell(current_row, mean_col)
                        sd_cell   = ws.cell(current_row, sd_col)
                        se_cell   = ws.cell(current_row, se_col)

                        mean_cell.value = f"=AVERAGE({tc}{r1}:{tc}{r3})"
                        sd_cell.value   = f"=STDEV({tc}{r1}:{tc}{r3})"
                        se_cell.value   = f"={sc}{r1}/SQRT(3)"

                        for stat_cell in (mean_cell, sd_cell, se_cell):
                            stat_cell.number_format = "0.00"
                            stat_cell.alignment     = CENTER
                    # Rep 2 & 3: mean/SD/SE cells remain empty

                    trait_col += 4

                current_row += 1

        # ---- Freeze panes: keep header visible -------------------------
        ws.freeze_panes = ws.cell(2, 1)

        # ---- Column widths ---------------------------------------------
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 5
        col = 3
        for _ in TRAIT_DISPLAY:
            ws.column_dimensions[get_column_letter(col)].width     = 18   # trait name
            ws.column_dimensions[get_column_letter(col + 1)].width = 8    # Mean
            ws.column_dimensions[get_column_letter(col + 2)].width = 8    # SD
            ws.column_dimensions[get_column_letter(col + 3)].width = 8    # SE
            col += 4

        data_rows = current_row - 2
        print(f"  [SHEET] {sheet_name:<10s}: {len(cultivars):2d} cultivars, {data_rows:3d} data rows")

    wb_out.save(str(output_path))
    print(f"\nSaved → {output_path}")
    print(f"Sheets written: {len(wb_out.sheetnames)}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    print(f"Input : {INPUT_PATH}")
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_PATH}")

    print("Reading raw data …")
    data = read_raw_data(INPUT_PATH)
    print(f"\nDates found: {len(data)}")

    print(f"\nOutput: {OUTPUT_PATH}")
    write_output(data, OUTPUT_PATH)


if __name__ == "__main__":
    main()
