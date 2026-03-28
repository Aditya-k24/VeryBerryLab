"""
src/etl.py
==========
Ingest Phenotyping 4 Worksheet 2.xlsx (wide cross-tab, 11 cultivar sheets)
and produce a tidy, row-per-replicate DataFrame with batch_id.

Missing-value semantics
-----------------------
  '-' cells    → NaN  (scheduled but not measured)
  absent dates → these rows simply don't exist (not_scheduled, handled in
                 aggregate.build_completeness_matrix via batch logic)

Run from the Analytics/ directory:
    python3 src/etl.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

ROOT     = Path(__file__).resolve().parent.parent          # Analytics/
XLSX     = ROOT / "data" / "Phenotyping 4 Worksheet 2.xlsx"
OUT_CSV  = ROOT / "data" / "processed" / "pheno4_clean.csv"
OUT_CSV.parent.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Lookup tables
# ---------------------------------------------------------------------------

CULTIVAR_MAP = {
    "Fin":   "Finn",
    "Cab":   "Cabrio",
    "Cam":   "Camarosa",   # provisional — confirm with team
    "Sen":   "Sensation",
    "Cha":   "Chandler",
    "Alb ":  "Albion",
    "Alb":   "Albion",
    "Mox":   "Moxie",
    "RJune": "Ruby June",
    "Bri":   "Brilliance",
    "Por":   "Portola",
    "Rad":   "Radiance",
}

BATCH_A = {"Albion", "Cabrio", "Camarosa", "Chandler", "Finn", "Sensation"}
BATCH_B = {"Brilliance", "Moxie", "Portola", "Radiance", "Ruby June"}

TRAIT_MAP = {
    "Pri stolon":                "n_stolon_primary",
    "Sec stolon":                "n_stolon_secondary",
    "Ter stolon":                "n_stolon_tertiary",
    "Quart Stolon":              "n_stolon_quaternary",
    "dp on alt of pri stolon":   "n_dp_alt_primary",
    "dp on alt of sec stolon":   "n_dp_alt_secondary",
    "dp on alt of ter stolon":   "n_dp_alt_tertiary",
    "dp on alt of quart stolon": "n_dp_alt_quaternary",
    "dp on mid of pri stolon":   "n_dp_mid_primary",
    "dp on mid of sec stolon":   "n_dp_mid_secondary",
    "dp on mid of ter stolon":   "n_dp_mid_tertiary",
    "dp on mid of quart stolon": "n_dp_mid_quaternary",
    "Total dp on alt":           "n_dp_total_alt",
    "Total dp on mid":           "n_dp_total_mid",
    "#Total Flowers":            "n_flowers_total",
    "# Flowers mp/mp":           "n_flowers_mp",
    "# Flowers dp/mp":           "n_flowers_dp",
    "Pri Stolon length (cm)":    "stolon_length_primary_cm",
    "Crown diameter (mm)":       "crown_diameter_mm",
}

TRAIT_COLS = list(TRAIT_MAP.values())

TRAIT_LABELS = {
    "n_stolon_primary":         "Primary Stolons",
    "n_stolon_secondary":       "Secondary Stolons",
    "n_stolon_tertiary":        "Tertiary Stolons",
    "n_stolon_quaternary":      "Quaternary Stolons",
    "n_dp_alt_primary":         "Daughter Plants (Alt) — Primary",
    "n_dp_alt_secondary":       "Daughter Plants (Alt) — Secondary",
    "n_dp_alt_tertiary":        "Daughter Plants (Alt) — Tertiary",
    "n_dp_alt_quaternary":      "Daughter Plants (Alt) — Quaternary",
    "n_dp_mid_primary":         "Daughter Plants (Mid) — Primary",
    "n_dp_mid_secondary":       "Daughter Plants (Mid) — Secondary",
    "n_dp_mid_tertiary":        "Daughter Plants (Mid) — Tertiary",
    "n_dp_mid_quaternary":      "Daughter Plants (Mid) — Quaternary",
    "n_dp_total_alt":           "Total DPs — Alternate Nodes",
    "n_dp_total_mid":           "Total DPs — Mid Nodes",
    "n_flowers_total":          "Total Flowers",
    "n_flowers_mp":             "Flowers on Mother Plant",
    "n_flowers_dp":             "Flowers on Daughter Plants",
    "stolon_length_primary_cm": "Primary Stolon Length (cm)",
    "crown_diameter_mm":        "Crown Diameter (mm)",
}

# True = higher is better for champion scoring
TRAIT_DIRECTION = {t: True for t in TRAIT_COLS}

ALL_CULTIVARS = sorted(BATCH_A | BATCH_B)


# ---------------------------------------------------------------------------
# Sheet parser
# ---------------------------------------------------------------------------

def _parse_sheet(ws, cultivar: str) -> tuple[list[dict], list[str]]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return [], [f"{cultivar}: fewer than 3 rows"]

    warnings: list[str] = []

    # Forward-fill dates (merged cells → None)
    dates: list = []
    cur = None
    for v in rows[0][1:]:
        if v is not None:
            cur = v
        dates.append(cur)

    reps: list = []
    for v in rows[1][1:]:
        try:
            reps.append(int(v))
        except (TypeError, ValueError):
            reps.append(None)

    col_headers = list(zip(dates, reps))

    # Skeleton: one record per (date, rep)
    skeleton: dict[tuple, dict] = {}
    for date, rep in col_headers:
        key = (date, rep)
        if key not in skeleton:
            skeleton[key] = {"cultivar": cultivar, "date": date, "rep": rep}

    # Fill trait values
    for row in rows[2:]:
        label = row[0]
        if label is None:
            continue
        internal = TRAIT_MAP.get(str(label).strip())
        if internal is None:
            continue
        for (date, rep), raw in zip(col_headers, row[1:]):
            if raw in ("-", "--") or raw is None:
                val = float("nan")
            else:
                try:
                    val = float(raw)
                except (TypeError, ValueError):
                    warnings.append(
                        f"{cultivar}: unexpected value '{raw}' for {internal} — treated as NaN"
                    )
                    val = float("nan")
            skeleton[(date, rep)][internal] = val

    return list(skeleton.values()), warnings


# ---------------------------------------------------------------------------
# Main ETL
# ---------------------------------------------------------------------------

def run_etl(verbose: bool = True) -> tuple[pd.DataFrame, list[str]]:
    """
    Parse XLSX → tidy DataFrame.
    Returns (df, ingestion_warnings).
    """
    if verbose:
        print(f"Loading: {XLSX}")

    wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)

    all_records: list[dict] = []
    all_warnings: list[str] = [
        "Formula results rely on cached Excel values. If you see unexpected blanks, "
        "open the file in Excel, recalculate, and re-save."
    ]

    for tab in wb.sheetnames:
        cv = CULTIVAR_MAP.get(tab)
        if cv is None:
            all_warnings.append(f"Unknown tab '{tab}' — skipped")
            continue
        records, warns = _parse_sheet(wb[tab], cv)
        all_records.extend(records)
        all_warnings.extend(warns)
        if verbose:
            print(f"  [OK]  {tab:8s} → {cv:12s}  ({len(records)} rows)")

    wb.close()

    df = _build_df(all_records)

    if verbose:
        _validate(df)
        df.to_csv(str(OUT_CSV), index=False)
        print(f"\nSaved → {OUT_CSV}  ({df.shape[0]} rows × {df.shape[1]} cols)")

    return df, all_warnings


def _build_df(records: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(records)

    for col in TRAIT_COLS:
        if col not in df.columns:
            df[col] = float("nan")

    df["date"] = pd.to_datetime(df["date"])
    df["rep"]  = pd.to_numeric(df["rep"], errors="coerce").astype("Int64")
    for col in TRAIT_COLS:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df["batch"] = df["cultivar"].apply(
        lambda cv: "A" if cv in BATCH_A else "B"
    )

    col_order = ["date", "cultivar", "batch", "rep"] + TRAIT_COLS
    df = df[[c for c in col_order if c in df.columns]]
    return df.sort_values(["cultivar", "date", "rep"]).reset_index(drop=True)


def _validate(df: pd.DataFrame) -> None:
    print("\n--- Validation ---")
    expected = set(ALL_CULTIVARS)
    found = set(df["cultivar"].unique())
    diff = expected.symmetric_difference(found)
    print(f"  {'[OK]  ' if not diff else '[WARN]'} Cultivars: {sorted(found)}")
    dupes = df.duplicated(["date", "cultivar", "rep"])
    print(f"  [{'OK  ' if not dupes.any() else 'FAIL'}] Duplicate rows: {dupes.sum()}")
    print(f"  [OK]  Date range: {df['date'].min().date()} → {df['date'].max().date()}")
    print(f"  [OK]  Shape: {df.shape}")


if __name__ == "__main__":
    df, warnings = run_etl()
    print("\nWarnings:")
    for w in warnings:
        print(f"  • {w}")
    print(df.head())
